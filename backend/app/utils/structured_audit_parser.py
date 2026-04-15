from __future__ import annotations

from datetime import datetime
import re
import unicodedata

from openpyxl import load_workbook

from app.models.audit_input import ApplicationScope, AuditObservation, MissionInfo, StructuredAuditInput


MISSION_HEADER_ALIASES = {
    "mission_id": {"id mission", "mission id", "mission_id"},
    "titre_mission": {"titre mission", "mission title", "titre_mission"},
    "type_mission": {"type mission", "mission type", "type_mission"},
    "entite_auditee": {"entite auditee", "entite auditee ", "audited entity", "entite_auditee"},
    "periode": {"periode", "period", "periode d audit"},
    "intervenants": {"intervenants", "stakeholders"},
    "perimetre_intervention": {"perimetre intervention", "perimetre d intervention", "scope", "perimetre_intervention"},
    "date_rapport": {"date rapport", "report date", "date_rapport"},
    "processus_couverts": {"processus couverts", "covered processes", "processus_couverts"},
    "objectifs": {"objectifs", "objectives"},
}

PERIMETER_HEADER_ALIASES = {
    "nom_application": {"nom application", "application", "application name", "nom_application"},
    "description": {"description", "description application", "description de l application", "description de l'application"},
    "systeme_exploitation": {"systeme d exploitation", "systeme d'exploitation", "systeme exploitation", "operating system", "os"},
    "base_donnees": {"base de donnees", "base de données", "database", "db", "sgbd"},
    "prestataire": {"prestataire", "provider", "vendor"},
}

OBSERVATION_HEADER_ALIASES = {
    "observation_id": {"id controle", "id observation", "observation id", "observation_id"},
    "domaine_controle": {"processus", "domaine controle", "domaine de controle", "control domain", "domaine_controle"},
    "categorie_controle": {"controle", "categorie controle", "categorie de controle", "control category", "categorie_controle"},
    "controle_ref": {"reference controle", "reference du controle", "controle ref", "control reference", "controle_ref"},
    "application": {"application", "systeme", "system"},
    "couche": {"couche", "layer"},
    "titre_observation": {"titre observation", "observation title", "titre_observation"},
    "constat": {"constat", "finding"},
    "risque_associe": {"risque", "risque associe", "risque associé", "risk", "risque_associe"},
    "procedure_compensatoire": {"procedure compensatoire", "compensating procedure", "procedure_compensatoire"},
    "cause_racine": {"cause racine", "root cause", "cause_racine"},
    "recommandation_proposee": {"recommandation", "recommandation proposee", "recommandation proposée", "proposed recommendation", "recommandation_proposee"},
    "commentaire_auditeur": {"commentaire", "commentaire auditeur", "auditor comment", "commentaire_auditeur"},
    "controle_attendu": {"controle attendu", "expected control", "controle_attendu"},
    "impact_potentiel": {"impact potentiel", "risk impact", "impact_potentiel"},
    "population": {"population", "population testee", "population testée", "tested population"},
    "taille_echantillon": {"taille echantillon", "taille échantillon", "sample size", "taille_echantillon"},
    "nombre_exceptions": {"nombre exceptions", "nb exceptions", "exceptions", "exception count", "nombre_exceptions"},
    "responsables": {"acteurs / responsables", "acteurs/responsables", "responsables", "owners", "owner"},
    "references_probantes": {"references probantes", "références probantes", "preuves", "evidence", "references_probantes"},
    "statut_controle": {"statut controle", "statut du controle", "résultat du test", "resultat du test", "control status", "statut_controle"},
    "statut_validation": {"statut validation", "validation status", "statut_validation"},
}


def _normalize(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def _normalize_header(value: str) -> str:
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    ascii_value = ascii_value.lower().strip()
    ascii_value = re.sub(r"[\n\r\t]+", " ", ascii_value)
    ascii_value = re.sub(r"[_\-]+", " ", ascii_value)
    ascii_value = re.sub(r"\s+", " ", ascii_value)
    return ascii_value


def _clean_text(value: str) -> str:
    text = value.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _split_values(value: str, separators: tuple[str, ...]) -> list[str]:
    normalized = _clean_text(value)
    for separator in separators[1:]:
        normalized = normalized.replace(separator, separators[0])
    return [item.strip() for item in normalized.split(separators[0]) if item.strip()]


def _match_header(raw_header: str, aliases: dict[str, set[str]]) -> str:
    normalized = _normalize_header(raw_header)
    for target, accepted_names in aliases.items():
        if normalized in accepted_names:
            return target
    return ""


def _find_header_row(rows: list[tuple], aliases: dict[str, set[str]], minimum_matches: int) -> tuple[int, list[str]]:
    for index, row in enumerate(rows):
        mapped_headers = [_match_header(_normalize(cell), aliases) for cell in row]
        if len([header for header in mapped_headers if header]) >= minimum_matches:
            return index, mapped_headers
    raise ValueError("Unable to identify the header row in the uploaded audit workbook.")


def _build_row_map(headers: list[str], values: list[str]) -> dict[str, str]:
    row_map: dict[str, str] = {}
    for header, value in zip(headers, values):
        if not header:
            continue
        row_map[header] = _clean_text(value)
    return row_map


def _find_sheet_name(workbook, *candidates: str) -> str:
    normalized_names = {_normalize_header(name): name for name in workbook.sheetnames}
    for candidate in candidates:
        matched = normalized_names.get(_normalize_header(candidate))
        if matched:
            return matched
    return ""


def parse_audit_workbook(file_path: str) -> StructuredAuditInput:
    workbook = load_workbook(file_path, data_only=True)

    mission_sheet = workbook["Mission"]
    mission_rows = list(mission_sheet.iter_rows(values_only=True))
    mission_header_index, mission_headers = _find_header_row(mission_rows, MISSION_HEADER_ALIASES, minimum_matches=5)

    mission_map: dict[str, str] = {}
    for row in mission_rows[mission_header_index + 1:]:
        values = [_normalize(cell) for cell in row]
        if not any(values):
            continue
        mission_map = _build_row_map(mission_headers, values)
        if mission_map:
            break

    perimeter_applications: list[str] = []
    perimeter_details: list[ApplicationScope] = []
    perimeter_sheet_name = _find_sheet_name(
        workbook,
        "Périmètre Intervention",
        "Perimetre Intervention",
        "Périmètre d'intervention",
        "Perimetre d'intervention",
        "Scope",
    )
    if perimeter_sheet_name:
        perimeter_sheet = workbook[perimeter_sheet_name]
        perimeter_rows = list(perimeter_sheet.iter_rows(values_only=True))
        try:
            perimeter_header_index, perimeter_headers = _find_header_row(
                perimeter_rows,
                PERIMETER_HEADER_ALIASES,
                minimum_matches=1,
            )
            for row in perimeter_rows[perimeter_header_index + 1 :]:
                values = [_normalize(cell) for cell in row]
                if not any(values):
                    continue
                row_map = _build_row_map(perimeter_headers, values)
                application_name = row_map.get("nom_application", "")
                if application_name:
                    perimeter_applications.append(application_name)
                    perimeter_details.append(
                        ApplicationScope(
                            name=application_name,
                            description=row_map.get("description", ""),
                            operating_system=row_map.get("systeme_exploitation", ""),
                            database=row_map.get("base_donnees", ""),
                            provider=row_map.get("prestataire", ""),
                        )
                    )
        except ValueError:
            perimeter_applications = []
            perimeter_details = []

    mission = MissionInfo(
        mission_id=mission_map.get("mission_id", ""),
        titre_mission=mission_map.get("titre_mission", ""),
        entite_auditee=mission_map.get("entite_auditee", ""),
        type_mission=mission_map.get("type_mission", ""),
        periode=mission_map.get("periode", ""),
        intervenants=_split_values(mission_map.get("intervenants", ""), (";", ",", "\n")),
        perimetre_intervention=mission_map.get("perimetre_intervention", ""),
        objectifs=_split_values(mission_map.get("objectifs", ""), (";", "\n")),
        date_rapport=mission_map.get("date_rapport", ""),
        processus_couverts=_split_values(mission_map.get("processus_couverts", ""), (";", ",", "\n")),
        applications=perimeter_applications or _split_values(mission_map.get("perimetre_intervention", ""), (";", ",", "\n")),
        application_details=perimeter_details,
    )

    observations_sheet = workbook["Observations"]
    observation_rows = list(observations_sheet.iter_rows(values_only=True))
    observation_header_index, observation_headers = _find_header_row(
        observation_rows,
        OBSERVATION_HEADER_ALIASES,
        minimum_matches=6,
    )

    observations: list[AuditObservation] = []
    for row in observation_rows[observation_header_index + 1:]:
        values = [_normalize(cell) for cell in row]
        if not any(values):
            continue
        row_map = _build_row_map(observation_headers, values)
        if not row_map.get("observation_id") and not row_map.get("titre_observation"):
            continue
        observations.append(AuditObservation(**row_map))

    return StructuredAuditInput(mission=mission, observations=observations)
