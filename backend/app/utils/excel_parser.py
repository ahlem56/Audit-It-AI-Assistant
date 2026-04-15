from __future__ import annotations

from pathlib import Path
from typing import Optional, Tuple

from langchain_core.documents import Document
from openpyxl import load_workbook

SKIPPED_SHEETS = {"guide"}


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_header_row(rows: list[tuple]) -> Optional[Tuple[int, list[str]]]:
    best_index = None
    best_headers: list[str] = []

    for index, row in enumerate(rows[:10]):
        candidate_headers = [_normalize_cell(cell) for cell in row]
        non_empty = [header for header in candidate_headers if header]
        if len(non_empty) > len([header for header in best_headers if header]):
            best_index = index
            best_headers = candidate_headers

    if best_index is None or not any(best_headers):
        return None
    return best_index, best_headers


def _sheet_rows_to_documents(workbook_path: str) -> list[Document]:
    workbook = load_workbook(workbook_path, data_only=True)
    documents: list[Document] = []

    for sheet in workbook.worksheets:
        if sheet.title.strip().lower() in SKIPPED_SHEETS:
            continue

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        header_info = _find_header_row(rows)
        if not header_info:
            continue

        header_row_index, raw_headers = header_info
        headers = [
            header or f"column_{index + 1}"
            for index, header in enumerate(raw_headers)
        ]

        for row_index, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
            values = [_normalize_cell(cell) for cell in row]
            row_data = {
                header: value
                for header, value in zip(headers, values)
                if value
            }
            if not row_data:
                continue

            content = "\n".join(f"{key}: {value}" for key, value in row_data.items())
            documents.append(
                Document(
                    page_content=content,
                    metadata={
                        "source_type": "excel",
                        "sheet_name": sheet.title,
                        "row_number": row_index,
                        "file_name": Path(workbook_path).name,
                    },
                )
            )

    return documents


def load_excel_document(file_path: str) -> list[Document]:
    return _sheet_rows_to_documents(file_path)
